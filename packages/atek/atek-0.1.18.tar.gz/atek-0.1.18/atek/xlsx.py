#!/usr/bin/env python
# -*- coding: utf-8 -*-

import typing as tp
import openpyxl
from pathlib import Path
import cytoolz.curried as tz
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font
import re
from collections import Counter
from dataclasses import dataclass, field, asdict, astuple
import atek
from fnmatch import fnmatch

__all__ = "fcell format_sheet add_table format_file sheet_columns".split()

def workbook(path: Path) -> Workbook:
    """Returns an Workbook object"""
    return tz.pipe(
        path
        ,atek.valid_path
        ,str
        ,openpyxl.load_workbook
    )


@dataclass(frozen=True)
class CellFormat:
    font: Font
    align: Alignment
    number_format: str
    column_width: int


class Format(tp.NamedTuple):
    pattern: str
    cell_format: CellFormat

    def validate(self):
        assert isinstance(self.cell_format, CellFormat), \
            "cell_format is not of type CellFormat"
        assert isinstance(self.pattern, str), "pattern must be a string"


def fcell(pattern: str, **kwargs) -> Format:
    """Creates default formats and merges them with provided formats.
    Used as a helper function to format_sheet."""
    # Set base format
    base = {"horizontal": "left", "vertical": "top", "wrap_text": True,
        "column_width": 20, "dtype": "", "number_format": ""}

    # Merge in provided args
    kwargs = kwargs or {}
    all_args = tz.merge(base, kwargs)

    # Calculate the dtype if not provided
    dtype = all_args["dtype"]
    num = all_args["number_format"].lower()
    right = {"horizontal": "right"}
    comma2 = {"number_format": "#,##0.00"}
    ymd = {"number_format": "yyyy-mm-dd"}
    extra = (
        right if dtype in ["n", "d"] and num != "" else
        tz.merge(right, comma2) if dtype == "n" and num == "" else
        tz.merge(right, ymd) if dtype == "d" and num == "" else
        right if "#" in num else
        right if "%" in num else
        right if "yy" in num else
        right if "m" in num else
        right if "d" in num else
        {}
    )

    all_args = tz.merge(all_args, extra)

    font_args = (
        "name bold italic strike outline shadow condense "
        "color extend size underline vertAlign"
    ).split()

    align_args = (
        "horizontal vertical text_rotation wrap_text shrink_to_fit "
        "indent"
    ).split()

    get_args = lambda arg_set: {
        k: v
        for k, v in all_args.items()
        if k in arg_set
    }

    return Format(
        pattern,
        CellFormat(
            font=Font(**get_args(font_args)),
            align=Alignment(**get_args(align_args)),
            number_format=all_args["number_format"],
            column_width=all_args["column_width"]
        )
    )


def match(dtype: str, column_name: str, *formats: Format) -> Format:
    for fmt in formats:
        # If normal tuple is provided then convert to Format
        _fmt = fmt if isinstance(fmt, Format) else Format(*fmt)
        _fmt.validate()

        # If match found return value
        if fnmatch(column_name, _fmt.pattern):
            return Format(column_name, _fmt.cell_format)

    # If no matches found, return default Format
    return fcell(column_name, dtype=dtype)


@dataclass(frozen=True)
class Column:
    column_name: str
    ref: str
    ref_header: str
    ref_data: str
    dtypes: Counter
    number_formats: Counter
    column_dtype: str
    column_desc: str
    max_width: int


@tz.curry
def columns_from_sheet(ws: Worksheet, rows: tp.Optional[int]=None) -> tp.List[Column]:

    desc = {
        "d": "date or datetime",
        "n": "number",
        "s": "text",
        "f": "formula",
        ".": "blank",
    }

    result = []

    for cells in ws.iter_cols(max_row=rows):
        column_name   = cells[0].value
        ref_header    = f"{cells[0].coordinate}"
        ref_data      = f"{cells[1].coordinate}:{cells[-1].coordinate}"
        ref           = f"{ref_header}:{cells[-1].coordinate}"
        dtypes        = Counter("." if cell.value == "" or cell.value is None else cell.data_type for cell in cells[1:])
        number_formats = Counter("." if cell.value == "" or cell.value is None else cell.number_format for cell in cells[1:])
        column_dtype  = dtypes.most_common(1)[0][0]
        max_width     = max(len(str(cell.value)) for cell in cells)
        column_desc   = desc.get(column_dtype, "unknown")

        column = Column(
            column_name,
            ref,
            ref_header,
            ref_data,
            dtypes,
            number_formats,
            column_dtype,
            column_desc,
            max_width
        )

        result.append(column)

    return result


@tz.curry
def sheet_columns(sheet_name: str, path: Path, rows: tp.Optional[int]=None) -> tp.List[Column]:

    wb = workbook(path)
    ws = wb[sheet_name]

    return columns_from_sheet(ws)


@tz.curry
def format_sheet(
    sheet_name: str,
    path: Path,
    *formats: Format,
    freeze_panes: str="") -> Path:

    wb = workbook(path)
    ws = wb[sheet_name]

    #  for column_index, cells in enumerate(ws.iter_cols(), start=1):
    for cells in ws.iter_cols():

        header = cells[0]
        column_name = header.value
        data = cells[1:]
        data_types = Counter(cell.data_type for cell in data)
        column_dtype = data_types.most_common(1)[0][0]
        max_width = max(len(str(cell.value)) for cell in cells)

        # Set freeze_panes for the first cell if it was not provided
        if freeze_panes == "":
            freeze_panes = data[0].coordinate

        # Get formats
        hfmt = fcell(column_name, dtype=column_dtype).cell_format
        dfmt = match(column_dtype, column_name, *formats).cell_format

        # Set column width
        width = min(max_width + 3, dfmt.column_width)
        ws.column_dimensions[header.column_letter].width = width
    #
        # Apply header format
        header.number_format = hfmt.number_format
        header.font = hfmt.font
        header.alignment = hfmt.align

        # Apply data formats
        for cell in data:
            cell.number_format = dfmt.number_format
            cell.font = dfmt.font
            cell.alignment = dfmt.align

    #  # Set the freeze_panes
    ws.freeze_panes = freeze_panes

    save_path = tz.pipe(path, atek.valid_path, str)
    wb.save(save_path)

    return save_path


@tz.curry
def table_name(sheet_name: str) -> str:
    new_name = re.sub("[\\\/\*\[\]:\? ]", "_", sheet_name)
    return new_name


@tz.curry
def add_table(sheet_name: str, path: Path, display_name: str="",
    showFirstColumn: bool=False, showLastColumn=False,
    showColumnStripes: bool=False, showRowStripes: bool=True,
    style_name: str="TableStyleLight15") -> Path:

    wb = workbook(path)
    ws = wb[sheet_name]

    data_range = ws.dimensions
    display_name = display_name or table_name(ws.title)

    table = Table(displayName=display_name, ref=data_range)
    style = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showColumnStripes=False,
        showRowStripes=True,
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    save_path = tz.pipe(path, atek.valid_path, str)
    wb.save(save_path)

    return save_path


def format_file(path: Path) -> Path:
    """Formats all sheets in a file using the default formats and add
    worksheet tables to each sheet."""
    wb = workbook(path)
    for sheet in wb.sheetnames:
        tz.pipe(
            path
            ,format_sheet(sheet)
            #  ,add_table(sheet)
        )

    return atek.valid_path(path)


if __name__ == "__main__":
    import pandas as pd
    path = Path.cwd() / "TestFile.xlsx"
    sheet="Just a test"

    sheet_info = lambda caption: tz.pipe(
        path
        ,sheet_columns(sheet)
        ,tz.map(asdict)
        ,atek.show(caption=caption)
    )

    export = lambda: tz.pipe(
        atek.query_atek02_main("select order_date, estvalue from orders limit 10")
        ,pd.DataFrame.from_records
        ,lambda df: df
        .to_excel(path, sheet_name=sheet, index=False)
    )

    print("TESTING format_sheet and add_table")
    export()
    sheet_info("Before formatting")

    tz.pipe(
        format_sheet(
            sheet,
            path,
            fcell("order_date", number_format="m/d/yyyy"),
            fcell("estvalue", number_format="$ #,##0.00"),
        )
        ,add_table
    )

    sheet_info("After formatting")

    print("TESTING format_file")
    export()
    sheet_info("Before formatting")
    format_file(path)
    sheet_info("After formatting")

    if path.exists:
        path.unlink()
