
import typing as tp
import datetime

import numpy as np

from static_frame.core.container_util import apex_to_name
from static_frame.core.doc_str import doc_inject
from static_frame.core.frame import Frame
from static_frame.core.index import Index
from static_frame.core.index_base import IndexBase
from static_frame.core.index_hierarchy import IndexHierarchy
from static_frame.core.store import Store
from static_frame.core.store import store_coherent_non_write
from static_frame.core.store import store_coherent_write
from static_frame.core.store import StoreConfig
from static_frame.core.store import StoreConfigMap
from static_frame.core.store import StoreConfigMapInitializer
from static_frame.core.store_filter import STORE_FILTER_DEFAULT
from static_frame.core.store_filter import StoreFilter
from static_frame.core.util import AnyCallable
from static_frame.core.util import BOOL_TYPES
from static_frame.core.util import COMPLEX_TYPES
# from static_frame.core.util import DT64_MONTH
# from static_frame.core.util import DT64_YEAR
from static_frame.core.util import DTYPE_BOOL
from static_frame.core.util import DTYPE_INEXACT_KINDS
from static_frame.core.util import DTYPE_INT_KINDS
# from static_frame.core.util import DTYPE_NAT_KINDS
from static_frame.core.util import DTYPE_OBJECT
from static_frame.core.util import DTYPE_STR_KINDS
from static_frame.core.util import NUMERIC_TYPES

if tp.TYPE_CHECKING:
    from xlsxwriter.worksheet import Worksheet  # pylint: disable=W0611 #pragma: no cover
    from xlsxwriter.workbook import Workbook  # pylint: disable=W0611 #pragma: no cover
    from xlsxwriter.format import Format  # pylint: disable=W0611 #pragma: no cover


MAX_XLSX_ROWS = 1048576
MAX_XLSX_COLUMNS = 16384 #1024 on libre office


class FormatDefaults:

    @staticmethod
    def label(f: 'Format') -> 'Format':
        f.set_bold()
        return f

    @staticmethod
    def date(f: 'Format') -> 'Format':
        f.set_num_format('yyyy-mm-dd')
        return f

    @staticmethod
    def datetime(f: 'Format') -> 'Format':
        f.set_num_format('yyyy-mm-ddThh:mm:ss.000') # ISO 8601 requires the T
        return f

    @staticmethod
    def get_format_or_default(
            workbook: 'Workbook', # do not want module level import o
            # format_specifier: tp.Optional[tp.Dict[str, tp.Any]],
            format_funcs: tp.Iterable[tp.Callable[['Format'], None]]
            ) -> 'Format':
        # if format_specifier:
        #     return workbook.add_format(format_specifier)
        f = workbook.add_format()
        for func in format_funcs:
            f = func(f)
        return f




class StoreXLSX(Store):

    _EXT: tp.FrozenSet[str] =  frozenset(('.xlsx',))

    # _EXT: str = '.xlsx'

    @staticmethod
    def _dtype_to_writer_attr(
            dtype: np.dtype,
            ) -> tp.Tuple[str, bool]:
        '''
        Return a pair of writer function, Boolean, where Boolean denotes if replacements need be applied.
        '''
        kind = dtype.kind

        # NOTE: xlsxwriter cannot handle datetime64, raises TypeError("Unknown or unsupported datetime type")
        # if kind in DTYPE_NAT_KINDS and dtype != DT64_MONTH and dtype != DT64_YEAR:
        #     return 'write_datetime', True

        if dtype == DTYPE_BOOL:
            return 'write_boolean', False
        elif kind in DTYPE_STR_KINDS:
            return 'write_string', False
        elif kind in DTYPE_INT_KINDS:
            return 'write_number', False
        elif kind in DTYPE_INEXACT_KINDS:
            return 'write_number', True
        return 'write', True

    @classmethod
    def _get_writer(cls,
            dtype: np.dtype,
            ws: 'Worksheet'
            ) -> AnyCallable: # find better type
        '''
        Return a writer function of the passed in Worksheet.
        '''
        assert isinstance(dtype, np.dtype)

        import xlsxwriter

        writer_attr, replace_active = cls._dtype_to_writer_attr(dtype)
        writer_native = getattr(ws, writer_attr)

        def writer(
                row: int,
                col: int,
                value: tp.Any,
                format_date: xlsxwriter.format.Format,
                format_datetime: xlsxwriter.format.Format,
                format_cell: tp.Optional[xlsxwriter.format.Format] = None,
                ) -> tp.Any:

            # cannot yet write complex types directly, so covert to string
            if isinstance(value, COMPLEX_TYPES):
                return ws.write_string(row, col, str(value), format_cell)

            if writer_attr == 'write':
                # determine type for each value
                if isinstance(value, BOOL_TYPES):
                    return ws.write_boolean(row, col, value, format_cell)
                if isinstance(value, str):
                    return ws.write_string(row, col, value, format_cell)
                if isinstance(value, NUMERIC_TYPES):
                    return ws.write_number(row, col, value, format_cell)
                if isinstance(value, datetime.datetime): # NOTE: must come before date isinstance check
                    return ws.write_datetime(row, col, value, format_datetime)
                if isinstance(value, datetime.date):
                    return ws.write_datetime(row, col, value, format_date)
            # use the type specific writer_native
            return writer_native(row, col, value, format_cell)
        return writer

    @classmethod
    def _frame_to_worksheet(cls,
            frame: Frame,
            ws: 'Worksheet',
            *,
            include_columns: bool,
            include_columns_name: bool = False,
            include_index: bool,
            include_index_name: bool = True,
            format_columns: 'Format',
            format_index: 'Format',
            format_date: 'Format',
            format_datetime: 'Format',
            format_columns_date: 'Format',
            format_columns_datetime: 'Format',
            format_index_date: 'Format',
            format_index_datetime: 'Format',
            merge_hierarchical_labels: bool,
            store_filter: tp.Optional[StoreFilter]
            ) -> None:

        if sum((include_columns_name, include_index_name)) > 1:
            raise RuntimeError('cannot set both `include_columns_name` and `include_index_name`')

        index_depth = frame._index.depth
        index_depth_effective = 0 if not include_index else index_depth
        index_names = frame._index.names # normalized presentation

        columns_iter = cls.get_column_iterator(frame=frame,
                include_index=include_index)

        columns_depth = frame._columns.depth
        columns_names = frame._columns.names
        columns_depth_effective = 0 if not include_columns else columns_depth

        columns_total = frame.shape[1] + index_depth_effective
        rows_total = frame.shape[0] + columns_depth_effective

        if rows_total > MAX_XLSX_ROWS:
            raise RuntimeError(f'Frame rows do not fit into XLSX sheet ({rows_total} > {MAX_XLSX_ROWS})')
        if columns_total > MAX_XLSX_COLUMNS:
            raise RuntimeError(f'Frame columns do not fit into XLSX sheet ({columns_total} > {MAX_XLSX_COLUMNS})')

        if include_columns:
            columns_values = frame._columns.values
            if store_filter:
                columns_values = store_filter.from_type_filter_array(columns_values)
            writer_columns = cls._get_writer(columns_values.dtype, ws)
            # for labels in acme, do not know type
            writer_names = cls._get_writer(DTYPE_OBJECT, ws)

        # write by column
        for col, values in enumerate(columns_iter):
            if include_columns:
                # The col integers will include index depth, so if including index, must wait until after index depth to write column field names; if include_index is False, can begin reading from columns_values
                if col < index_depth_effective:
                    if include_index_name:
                        writer_names(0, # always populate in top-most row
                                col,
                                index_names[col],
                                format_cell=format_index,
                                format_date=format_index_date,
                                format_datetime=format_index_datetime,
                                )
                    if include_columns_name and col == 0:
                        for i in range(columns_depth):
                            writer_names(i,
                                    col, # always 0, populate in left-most colum
                                    columns_names[i],
                                    format_cell=format_columns,
                                    format_date=format_columns_date,
                                    format_datetime=format_columns_datetime,
                                    )
                else: # col >= index_depth_effective:
                    if columns_depth == 1:
                        writer_columns(0,
                                col,
                                columns_values[col - index_depth_effective],
                                format_cell=format_columns,
                                format_date=format_columns_date,
                                format_datetime=format_columns_datetime,
                                )
                    elif columns_depth > 1:
                        for i in range(columns_depth):
                            # here, row selection is column count, column selection is depth
                            writer_columns(i,
                                    col,
                                    columns_values[col - index_depth_effective, i],
                                    format_cell=format_columns,
                                    format_date=format_columns_date,
                                    format_datetime=format_columns_datetime,
                                    )
            if store_filter:
                # thi might change the dtype
                values = store_filter.from_type_filter_array(values)
            writer = cls._get_writer(values.dtype, ws)
            # start enumeration of row after the effective column depth
            for row, v in enumerate(values, columns_depth_effective):
                if col < index_depth_effective:
                    writer(row,
                            col,
                            v,
                            format_cell=format_index,
                            format_date=format_index_date,
                            format_datetime=format_index_datetime,
                            )
                else:
                    writer(row,
                            col,
                            v,
                            format_cell=None,
                            format_date=format_date,
                            format_datetime=format_datetime,
                            )

        # post process to merge cells; need to get width of at depth
        if include_columns and merge_hierarchical_labels and columns_depth > 1:
            for depth in range(columns_depth - 1): # never most deep
                row = depth
                col = index_depth_effective # start after index
                for label, width in frame._columns.label_widths_at_depth(depth):
                    # TODO: use store_filter
                    ws.merge_range(row, col, row, col + width - 1, label, format_columns)
                    col += width

        if include_index and merge_hierarchical_labels and index_depth > 1:
            for depth in range(index_depth - 1): # never most deep
                row = columns_depth_effective
                col = depth
                for label, width in frame._index.label_widths_at_depth(depth):
                    # TODO: use store_filter
                    ws.merge_range(row, col, row + width - 1, col, label, format_index)
                    row += width

    @store_coherent_write
    def write(self,
            items: tp.Iterable[tp.Tuple[tp.Optional[str], Frame]],
            *,
            config: StoreConfigMapInitializer = None,
            store_filter: tp.Optional[StoreFilter] = STORE_FILTER_DEFAULT
            ) -> None:
        '''
        Args:
            store_filter: a dictionary of objects to string, enabling replacement of NaN and None values when writng to XLSX.

        '''
        # format_data: tp.Optional[tp.Dict[tp.Hashable, tp.Dict[str, tp.Any]]]
        # format_data: dictionary of dictionaries, keyed by column label, that contains dictionaries of XlsxWriter format specifications.

        # will create default from None, will pass let a map pass through
        config_map = StoreConfigMap.from_initializer(config)

        import xlsxwriter

        # NOTE: can supply second argument: {'default_date_format': 'dd/mm/yy'}
        wb = xlsxwriter.Workbook(self._fp, {'remove_timezone': True})

        for label, frame in items:
            c = config_map[label]

            # NOTE: this must be called here, as we need the workbook been assigning formats, and we need to get a config per label
            format_columns = FormatDefaults.get_format_or_default(
                    wb,
                    format_funcs=(FormatDefaults.label,))
            format_index = FormatDefaults.get_format_or_default(
                    wb,
                    format_funcs=(FormatDefaults.label,))

            format_date = FormatDefaults.get_format_or_default(
                    wb,
                    format_funcs=(FormatDefaults.date,))
            format_datetime = FormatDefaults.get_format_or_default(
                    wb,
                    format_funcs=(FormatDefaults.datetime,))

            format_columns_date = FormatDefaults.get_format_or_default(
                    wb,
                    format_funcs=(FormatDefaults.label, FormatDefaults.date))
            format_columns_datetime = FormatDefaults.get_format_or_default(
                    wb,
                    format_funcs=(FormatDefaults.label, FormatDefaults.datetime,))

            format_index_date = FormatDefaults.get_format_or_default(
                    wb,
                    format_funcs=(FormatDefaults.label, FormatDefaults.date))
            format_index_datetime = FormatDefaults.get_format_or_default(
                    wb,
                    format_funcs=(FormatDefaults.label, FormatDefaults.datetime,))


            ws = wb.add_worksheet(label)
            self._frame_to_worksheet(frame,
                    ws,
                    format_columns=format_columns,
                    format_index=format_index,
                    format_date=format_date,
                    format_datetime=format_datetime,
                    format_columns_date=format_columns_date,
                    format_columns_datetime=format_columns_datetime,
                    format_index_date=format_index_date,
                    format_index_datetime=format_index_datetime,
                    include_index=c.include_index,
                    include_index_name=c.include_index_name,
                    include_columns=c.include_columns,
                    include_columns_name=c.include_columns_name,
                    merge_hierarchical_labels=c.merge_hierarchical_labels,
                    store_filter=store_filter
                    )
        wb.close()

    @staticmethod
    def _load_workbook(fp: str) -> 'Workbook':
        import openpyxl
        return openpyxl.load_workbook(
                filename=fp,
                read_only=True,
                data_only=True
                )

    @doc_inject(selector='constructor_frame')
    @store_coherent_non_write
    def read(self,
            label: tp.Optional[str] = None,
            *,
            config: tp.Optional[StoreConfig] = None,
            store_filter: tp.Optional[StoreFilter] = STORE_FILTER_DEFAULT,
            container_type: tp.Type[Frame] = Frame,
            ) -> Frame:
        '''
        Args:
            label: Name of sheet to read from XLSX.
            container_type: Type of container to be returned, either Frame or a Frame subclass

        '''
        if config is None:
            config = StoreConfig() # get default

        index_depth = config.index_depth
        index_name_depth_level = config.index_name_depth_level
        columns_depth = config.columns_depth
        columns_name_depth_level = config.columns_name_depth_level

        skip_header = config.skip_header
        skip_footer = config.skip_footer

        wb = self._load_workbook(self._fp)

        if label is None:
            ws = wb[wb.sheetnames[0]]
            name = None # do not set to default sheet name
        else:
            ws = wb[label]
            name = ws.title

        if ws.max_column <= 1 or ws.max_row <= 1:
            # https://openpyxl.readthedocs.io/en/stable/optimized.html
            # says that some clients might not report correct dimensions
            ws.calculate_dimension()

        max_column = ws.max_column
        max_row = ws.max_row

        # adjust for downward shift for skipping header, then reduce for footer; at this value and beyond we stop
        last_row_count = max_row - skip_header - skip_footer

        index_values: tp.List[tp.Any] = []
        columns_values: tp.List[tp.Any] = []

        data = [] # pre-size with None?
        apex_rows = []

        for row_count, row in enumerate(ws.iter_rows(max_row=max_row), start=-skip_header):
            if row_count < 0:
                continue # due to skip header; perserves comparison to columns_depth
            if row_count >= last_row_count:
                break

            if store_filter is None:
                row = tuple(c.value for c in row)
            else: # only need to filter string values, but probably too expensive to pre-check
                row = tuple(store_filter.to_type_filter_element(c.value) for c in row)

            if row_count <= columns_depth - 1:
                apex_rows.append(row[:index_depth])
                if columns_depth == 1:
                    columns_values.extend(row[index_depth:])
                elif columns_depth > 1:
                    # NOTE: this orientation will need to be rotated
                    columns_values.append(row[index_depth:])
                continue

            if index_depth == 0:
                data.append(row)
            elif index_depth == 1:
                index_values.append(row[0])
                data.append(row[1:])
            else:
                index_values.append(row[:index_depth])
                data.append(row[index_depth:])

        wb.close()

        # Trim all-empty trailing rows created from style formatting GH#146. As the wb is opened in read-only mode, reverse iterating on the wb is not an option, nor is direct row access by integer; alos, evaluating all rows on forward iteration is expensive. Instead, after collecting all the data in a list and closing the wb, reverse iterate and find rows that are all empty.
        # NOTE: need to handle case where there are valid index values

        empty_token = (None if store_filter is None
                else store_filter.to_type_filter_element(None))

        for row_count in range(len(data) - 1, -2, -1):
            if row_count < 0:
                break
            if any(c != empty_token for c in data[row_count]): # try to break early with any
                break
            if index_depth == 1 and index_values[row_count] != empty_token:
                break
            if index_depth > 1 and any(c != empty_token for c in index_values[row_count]):
                break

        # row_count is set to the first row that has data or index; can be -1
        empty_row_idx = row_count + 1 # index of all-empty row
        if empty_row_idx != len(data):
            # trim data and index_values, if index_depth > 0
            data = data[:empty_row_idx]
            if index_depth > 0:
                index_values = index_values[:empty_row_idx]

        # continue with Index and Frame creation
        index_name = None if columns_depth == 0 else apex_to_name(
                rows=apex_rows,
                depth_level=index_name_depth_level,
                axis=0,
                axis_depth=index_depth)

        index: tp.Optional[IndexBase] = None
        own_index = False
        if index_depth == 1:
            index = Index(index_values, name=index_name)
            own_index = True
        elif index_depth > 1:
            index = IndexHierarchy.from_labels(
                    index_values,
                    continuation_token=None,
                    name=index_name,
                    )
            own_index = True

        columns_name = None if index_depth == 0 else apex_to_name(
                    rows=apex_rows,
                    depth_level=columns_name_depth_level,
                    axis=1,
                    axis_depth=columns_depth)

        columns: tp.Optional[IndexBase] = None
        own_columns = False
        if columns_depth == 1:
            columns = container_type._COLUMNS_CONSTRUCTOR(
                    columns_values,
                    name=columns_name)
            own_columns = True
        elif columns_depth > 1:
            columns = container_type._COLUMNS_HIERARCHY_CONSTRUCTOR.from_labels(
                    zip(*columns_values),
                    continuation_token=None,
                    name=columns_name,
                    )
            own_columns = True

        # NOTE: this might be a Frame or a FrameGO
        return container_type.from_records(data, #type: ignore
                        index=index,
                        columns=columns,
                        dtypes=config.dtypes,
                        own_index=own_index,
                        own_columns=own_columns,
                        name=name,
                        consolidate_blocks=config.consolidate_blocks
                        )

    @store_coherent_non_write
    def labels(self, strip_ext: bool = True) -> tp.Iterator[str]:
        wb = self._load_workbook(self._fp)
        labels = tuple(wb.sheetnames)
        wb.close()
        yield from labels



