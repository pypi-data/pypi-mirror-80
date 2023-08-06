# -*- coding: utf-8; -*-
################################################################################
#
#  Rattail -- Retail Software Framework
#  Copyright © 2010-2019 Lance Edgar
#
#  This file is part of Rattail.
#
#  Rattail is free software: you can redistribute it and/or modify it under the
#  terms of the GNU General Public License as published by the Free Software
#  Foundation, either version 3 of the License, or (at your option) any later
#  version.
#
#  Rattail is distributed in the hope that it will be useful, but WITHOUT ANY
#  WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS
#  FOR A PARTICULAR PURPOSE.  See the GNU General Public License for more
#  details.
#
#  You should have received a copy of the GNU General Public License along with
#  Rattail.  If not, see <http://www.gnu.org/licenses/>.
#
################################################################################
"""
Report Definitions
"""

from __future__ import unicode_literals, absolute_import

from openpyxl.styles import PatternFill

from rattail.reporting import Report
from rattail.excel import ExcelWriter


class ExcelReport(Report):
    """
    Generic report which knows how to write Excel output file.

    .. attr:: output_fields

       Simple list of field names which should be included in the output file.

    .. attr:: number_formats

       Optional dictionary specifying "number formats" for any fields.  Use the
       field name for dict key, and value should be the Excel-specific number
       format to be applied to all that column's values.

    .. attr:: totalled_fields

       Optional list of fields which should be "totalled" and represented in a
       final totals row within the output.
    """
    output_fields = []
    number_formats = {}
    totalled_fields = []

    def make_filename(self, session, params, data, **kwargs):
        return "{}.xlsx".format(self.name)

    def write_file(self, session, params, data, path, progress=None, **kwargs):
        """
        Write a basic Excel output file with the given data.  Requires at least
        the ``output_fields`` attribute to be set to work correctly.
        """
        writer = ExcelWriter(path, self.output_fields,
                             number_formats=self.number_formats)
        writer.write_header()

        # convert data to Excel-compatible rows
        rows = [
            [row[field] for field in self.output_fields]
            for row in data]

        # write main data rows
        writer.write_rows(rows, progress=progress)

        # maybe add a TOTALS row
        totals = {}
        for field in self.totalled_fields:
            totals[field] = sum([row[field] for row in data])
        if totals:

            # create totals row data
            rowdata = []
            for field in self.output_fields:
                if field in totals:
                    rowdata.append(totals[field])
                else:
                    rowdata.append(None)

            # append row to output
            writer.write_row(rowdata, row=len(data) + 2)

            # apply row highlighting
            fill_totals = PatternFill(patternType='solid',
                                      fgColor='ffee88',
                                      bgColor='ffee88')
            for col, field in enumerate(self.output_fields, 1):
                cell = writer.sheet.cell(row=len(data) + 2, column=col)
                cell.fill = fill_totals

        writer.auto_freeze()
        writer.auto_filter()
        writer.auto_resize()
        writer.save()
