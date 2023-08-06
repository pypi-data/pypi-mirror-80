"""
SheetTable Class: A class for query, update a exists Excel table.
"""
import os
import openpyxl
import shutil
from xlsxwriter.utility import xl_range


class SheetTable(object):
    def __init__(self, template_path, output_path, sheet_name, table_name):
        self.template_path = template_path
        self.output_path = output_path
        self.sheet_name = sheet_name
        self.table_name = table_name
        if self.template_path != self.output_path:
            # generate a new file at output path
            if os.path.exists(output_path):
                os.remove(output_path)
            shutil.copy(self.template_path, self.output_path)
        print("start load")
        self.workbook = openpyxl.load_workbook(self.output_path)
        print("end load")
        self.worksheet = self.workbook[sheet_name]

    def get_col_meta_data(self):
        nrow, ncol = self.get_shape()
        if nrow == 1:
            row_loc = 2
        else:
            row_loc = nrow
        col_meta_data = {}
        for col_loc in range(1, ncol + 1):
            num_format = self.worksheet.cell(row_loc, col_loc).number_format
            default_value = self.worksheet.cell(row_loc, col_loc).value
            default_value = (
                None
                if (default_value is None) or (len(default_value) == 0)
                else default_value
            )
            col_meta_data.update(
                {col_loc: {"num_format": num_format, "default_value": default_value}}
            )
        return col_meta_data

    def insert_records(self, records, auto_update_ref=True):
        nrow, ncol = self.get_shape()
        row_loc = nrow
        col_meta_data = self.get_col_meta_data()
        for record in records:
            assert len(record) == ncol
            for col_loc in range(1, ncol + 1):
                i = col_loc - 1
                num_format = col_meta_data.get(col_loc).get("num_format")
                default_value = col_meta_data.get(col_loc).get("default_value")
                if default_value is None:
                    record_value = record[i]
                else:
                    record_value = default_value
                # converted_record_value = record_value.get_formula_string(col_loc, row_loc)
                # if isinstance(record_value, CellFormula):
                #     converted_record_value = record_value.get_formula_string(
                #         col_loc, row_loc
                #     )
                # else:
                #     converted_record_value = record_value
                self.worksheet.cell(row_loc, col_loc).value = record_value
                if num_format is not None:
                    self.worksheet.cell(row_loc, col_loc).number_format = num_format
            row_loc += 1
        self.save()
        if auto_update_ref:
            self._update_ref()

    def _update_ref(self):
        table = self.worksheet.tables[self.table_name]
        ref = self.get_ref()
        table.ref = ref
        self.save()

    def get_shape(self):
        """获得包括字段名在内的表格行数和列数"""
        nrow = 0
        for row in self.worksheet.iter_rows():
            nrow += 1
        ncol = len(row)
        return (nrow, ncol)

    def get_rows(self):
        return [[cell.value for cell in row] for row in self.ws.iter_rows()]

    def get_ref(self):
        table_shape = self.get_shape()
        # 至少有一行的数值行
        final_cell_loc = (max([table_shape[0] - 1, 1]), table_shape[1] - 1)
        # xlsxwriter使用从零开始的索引
        ref = xl_range(0, 0, *final_cell_loc)
        return ref

    def save(self):
        self.workbook.save(self.output_path)
        self = self.__init__(
            self.output_path, self.output_path, self.sheet_name, self.table_name
        )


class CellFormula:
    def __init__(self, template):
        self.template = template

    def get_formula_string(self, col, row):
        formula_string = self.template.format(col=col, row=row)
        return formula_string
