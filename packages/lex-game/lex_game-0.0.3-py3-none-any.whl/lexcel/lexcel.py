import warnings
import logging
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import NamedStyle, Border, Side, Alignment, PatternFill
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle

from . import lex

class LExcel:

    CODE_ROW = 6 # Row of board code
    CODE_COL = 1 # Column of the first board code
    MOVES_ROW = 8 # Row of the first move
    MOVES_OFFSET = 0 # Offset of moves w.r.t. CODE_COL

    STYLE = NamedStyle(name="default")
    BORDER = Side(style='thin', color="000000")
    STYLE.border = Border(left=BORDER, top=BORDER, right=BORDER, bottom=BORDER)
    STYLE.alignment = Alignment(horizontal="center", vertical="center")
    CYAN_STYLE = Rule(type='cellIs', operator='equal',
                      dxf=DifferentialStyle(fill=PatternFill(bgColor='00FFFF')),
                      formula=["{}".format(lex.CYAN)])
    VIOLET_STYLE = Rule(type='cellIs', operator='equal',
                        dxf=DifferentialStyle(fill=PatternFill(bgColor='EE82EE')),
                        formula=["{}".format(lex.VIOLET)])

    def __init__(self, filename):
        self.FILE = filename
        warnings.simplefilter("ignore")
        self.WB = openpyxl.load_workbook(self.FILE, read_only=True, data_only=True)


    def find_board_column(self, turn, code):
        """Return the column of the board identified by code and turn.

        None if code is not found.
        """
        ws = self.WB['Turn{}'.format(turn)]
        ws.calculate_dimension(force=True) # Google sheets save as unsized
        for col in range(self.CODE_COL, ws.max_column+1):
            c = ws.cell(row=self.CODE_ROW, column=col)
            if c.value is not None:
                if c.value == code:
                    return c.column
        return None


    def get_moves(self, board, turn, player=lex.VIOLET):
        """Get the moves associated to a given board and turn.

        Return a pair of lists: the possible moves and the current weights,
        None if the board is not present in filename.
        Illegal moves are always implicitly weighted 0, thus they are never returned, rewarded or punished.
        """
        col = self.find_board_column(turn, board.get_code())
        if col is not None:
            moves_col = col + self.MOVES_OFFSET
            moves, weights = [], []
            ws = self.WB['Turn{}'.format(turn)]
            for r in range(self.MOVES_ROW, self.MOVES_ROW+len(board.WELL_FORMED_MOVES)+1):
                move_cell = ws[get_column_letter(moves_col)+str(r)]
                weight_cell = ws[get_column_letter(moves_col+1)+str(r)]
                if move_cell.value is not None:
                    legal_moves = board.get_legal_moves(player)
                    if move_cell.value in legal_moves:
                        moves.append(move_cell.value)
                        assert weight_cell.value is not None, "None weight {} {}".format(board.get_code(), turn) # pragma: no mutate
                        weights.append(weight_cell.value)
                    else:
                        moves.append(move_cell.value)
                        weights.append(0)

            return moves, weights
        return None

    def reward(self, newfile, match_moves, prize):
        """Write a new Excel file with every move in match_moves rewarded.
        """
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(self.FILE, data_only=True)
        for move, code, turn in match_moves:
            ws = wb['Turn{}'.format(turn)]
            col = self.find_board_column(turn, code)
            board = lex.Board(code)
            if col is not None:
                moves_col = col + self.MOVES_OFFSET
                for r in range(self.MOVES_ROW, self.MOVES_ROW+len(board.WELL_FORMED_MOVES)+1):
                    move_cell = ws[get_column_letter(moves_col)+str(r)]
                    weight_cell = ws[get_column_letter(moves_col+1)+str(r)]
                    if move_cell.value == move:
                        weight_cell.value = weight_cell.value + prize
                        break
        wb.save(newfile)
        wb.close()


    def choose_and_remember_move(self, board, turn, match_moves, rand, player=lex.VIOLET):
        """Randomly choose a move, append info in match_moves to reward in case of a win.

        """

        def choose_move(b):
            moves_weights = self.get_moves(b, turn, player)
            if moves_weights is not None:
                moves, weights = moves_weights
                if sum(weights) > 0:
                    m = rand.choices(moves, weights)
                    logging.debug('Choices: {} {} {}'.format(moves, weights, m)) # pragma: no mutate
                else:
                    m = moves[0:1]
                return m[0]
            else:
                return None

        move = choose_move(board)
        if move is not None:
            match_moves.append((move, board.get_code(), turn))
            return move

        move = choose_move(board.flip())
        if move is not None:
            match_moves.append((move, board.flip().get_code(), turn))
            return board.flip_move(move)

    @classmethod
    def make_empty_experience(cls, ncols, filename='exapawn-empty.xls', players=[lex.VIOLET, lex.CYAN]):
        board = lex.Board(ncols=ncols)
        _, boards = board.get_forest(len(board.COLS))
        wb = openpyxl.Workbook()
        for player in players:
            maxturn = 0
            for t, _ in boards.values():
                if t > maxturn:
                    maxturn = t
            for turn in range(player, maxturn+1, 2):
                name = 'Turn{}'.format(turn)
                ws = wb.create_sheet(name)
                code_col = cls.CODE_COL
                for code in sorted(boards): # sort values for reproducibility
                    t, moves = boards[code]
                    if t == turn:
                        board = lex.Board(code)
                        for i, c in enumerate(board.COLS):
                            col = code_col + i
                            x = ws.cell(row=1, column=col, value=c)
                            x.style = cls.STYLE
                            for r in range(2, 2+len(board[c])):
                                x = ws.cell(row=r, column=col, value=board[c][r-2])
                                x.style = cls.STYLE
                        x = ws.cell(row=cls.CODE_ROW, column=code_col, value=code)
                        x.style = cls.STYLE

                        moves_row=cls.MOVES_ROW
                        for m in sorted(moves): # sort values for reproducibility
                            x = ws.cell(row=moves_row, column=code_col+cls.MOVES_OFFSET, value=m)
                            x.style = cls.STYLE
                            x = ws.cell(row=moves_row, column=code_col+cls.MOVES_OFFSET+1, value=1)
                            x.style = cls.STYLE
                            moves_row = moves_row + 1

                        code_col = code_col + len(board.COLS) + 1
                ws.conditional_formatting.add('A2:{}4'.format(get_column_letter(code_col)), cls.CYAN_STYLE)
                ws.conditional_formatting.add('A2:{}4'.format(get_column_letter(code_col)), cls.VIOLET_STYLE)
        del wb['Sheet']
        wb.save(filename)
